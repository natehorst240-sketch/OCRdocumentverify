"""Export the compliance gap analysis to a color-coded Excel workbook (US 4.3).

Three tabs — Complied, Outstanding, Needs Review — each row carrying the
requirement source, description, compliance date/hours, confidence, and the
source document/page. The file is named with today's date and the tail number.
"""

from datetime import date
from pathlib import Path

import database

COLUMNS = [
    ("requirement_source", "Requirement Source"),
    ("doc_number", "Doc Number"),
    ("req_type", "Type"),
    ("description", "Description"),
    ("compliance_date", "Compliance Date"),
    ("compliance_hours", "Compliance Hours"),
    ("confidence", "Confidence"),
    ("record_source", "Source Document"),
    ("page_number", "Page"),
]

# Tab name -> (compliance statuses included, header fill color).
TABS = {
    "Complied": (("complied",), "C6EFCE"),        # green
    "Outstanding": (("outstanding", None), "FFC7CE"),  # red (None = no result)
    "Needs Review": (("needs_review",), "FFEB9C"),     # amber
}


def _safe_tail(tail_number: str) -> str:
    cleaned = "".join(c for c in (tail_number or "") if c.isalnum())
    return cleaned or "UNKNOWN"


def build_gap_report(tail_number: str, out_dir: str | Path,
                     db_path=database.DB_PATH) -> Path:
    """Write the gap-report workbook and return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = database.compliance_report(db_path)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    for tab_name, (statuses, color) in TABS.items():
        ws = wb.create_sheet(tab_name)
        header_fill = PatternFill("solid", fgColor=color)
        bold = Font(bold=True)

        for col_idx, (_, label) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = bold
            cell.fill = header_fill

        row_fill = PatternFill("solid", fgColor=color)
        out_row = 2
        for row in rows:
            if row["status"] not in statuses:
                continue
            for col_idx, (key, _) in enumerate(COLUMNS, start=1):
                value = row[key] if key in row.keys() else None
                cell = ws.cell(row=out_row, column=col_idx, value=value)
                cell.fill = row_fill
            out_row += 1

        # Reasonable column widths.
        for col_idx, (_, label) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx)
                                 .column_letter].width = max(14, len(label) + 2)
        ws.freeze_panes = "A2"

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = f"gap_report_{_safe_tail(tail_number)}_{date.today():%Y%m%d}.xlsx"
    out_path = out_dir / filename
    wb.save(out_path)
    return out_path
